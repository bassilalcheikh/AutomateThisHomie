# LenderX ETL
#
# Guaranteed Rate, Inc.
#
import openpyxl
import time
time_0 = time.time()
wb = openpyxl.load_workbook('C:\Python27\PythonFiles\AppraiserPerformance.xlsx') # runtime: ~0.141 seconds
time_1 = time.time()
main_sheet = wb.get_sheet_by_name('AppraiserPerformance') #spreadsheet object

#print(main_sheet['A2'].value) HEADS UP: seemingly "string" values are in unicode, seemingly "int" values are long
cell_A1 = main_sheet.cell(row=1, column =1)

print(main_sheet.max_column)
